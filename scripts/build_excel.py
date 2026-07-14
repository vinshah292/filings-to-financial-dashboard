"""
build_excel.py, the auditable Excel hub.

Reads the verified dashboard_data.json (+ quarterly detail) and writes a clean workbook that ties
EXACTLY to the web dashboard: same SEC-sourced numbers, same corrected net-income attribution.
Pulled figures are blue (inputs); ratios/per-MD/segment-totals are live Excel formulas (black), so
the model recomputes if an input changes. Gridlines off (house rule). Part of refresh.py.

Tabs: Cover · Operating Model (annual FY22-FY26) · Quarterly Detail (20 quarters, the audit trail).
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJ=r"C:\Users\vinsh\OneDrive\Documents\Claude\Projects\Github Finance Outputs\Houlihan Lokey Trend Analysis Project"
DATA=os.path.join(PROJ,"data")
D=json.load(open(os.path.join(PROJ,"dashboard_data.json"),encoding="utf-8"))
QF=json.load(open(os.path.join(DATA,"hli_qfin.json"),encoding="utf-8"))
QS=json.load(open(os.path.join(DATA,"hli_quarterly.json"),encoding="utf-8"))
FYS=D["meta"]["fys"]; s=lambda fy:str(fy)

NAVY="004878"; BLUE="1F6FA8"; INK="22333B"; GREY="5A6B72"; STEEL="5490A8"
f_h1   =Font(name="Calibri",size=16,bold=True,color=NAVY)
f_sub  =Font(name="Calibri",size=10,color=GREY)
f_secth=Font(name="Calibri",size=10,bold=True,color="FFFFFF")
f_lbl  =Font(name="Calibri",size=10,color=INK)
f_lblb =Font(name="Calibri",size=10,bold=True,color=INK)
f_input=Font(name="Calibri",size=10,color=BLUE)          # pulled from SEC
f_formu=Font(name="Calibri",size=10,color=INK)           # computed in-sheet
f_colh =Font(name="Calibri",size=10,bold=True,color="FFFFFF")
fill_navy=PatternFill("solid",fgColor=NAVY)
fill_sec =PatternFill("solid",fgColor=STEEL)
thin=Side(style="thin",color="D6DEE3")
btm =Border(bottom=thin)
R=Alignment(horizontal="right"); L=Alignment(horizontal="left"); C=Alignment(horizontal="center")

def sheet(wb,title):
    ws=wb.create_sheet(title); ws.sheet_view.showGridLines=False; return ws

wb=openpyxl.Workbook()

# ============================== COVER ==============================
cov=wb.active; cov.title="Cover"; cov.sheet_view.showGridLines=False
cov.column_dimensions["A"].width=2
cov.column_dimensions["B"].width=98
def cline(r,txt,font,fill=None):
    c=cov.cell(row=r,column=2,value=txt); c.font=font
    if fill: c.fill=fill; c.font=Font(name="Calibri",size=c.font.size,bold=c.font.bold,color="FFFFFF")
cline(2,"Houlihan Lokey, Inc.  (NYSE: HLI)",f_h1)
cline(3,"Five-Year Operating Model  ·  FY2022–FY2026  ·  fiscal years ending March 31",f_sub)
cov.cell(row=5,column=2,value="AUDITABLE DATA HUB").font=f_secth
cov.cell(row=5,column=2).fill=fill_navy
notes=[
 "",
 "Source of truth: SEC EDGAR. Financial lines are XBRL company facts (CIK 0001302215), YTD-differenced",
 "to quarters. Segment revenue, Managing Director headcount, and deal counts are parsed from the 20",
 "quarterly earnings-release 8-Ks (Exhibit 99.1).",
 "",
 "Every figure here ties to the web dashboard and to the filings. Blue cells are pulled from SEC; black",
 "cells are live Excel formulas (margins, comp ratio, revenue per MD, segment totals) that recompute.",
 "",
 "Tie-outs (enforced in code before this file is written):",
 "   • Segment revenue sums to total revenue, every fiscal year.",
 "   • Quarterly revenue sums to the reported 10-K annual, every fiscal year.",
 "",
 "Net income to HLI = us-gaap:NetIncomeLoss (attributable to parent), not consolidated ProfitLoss.",
 "EBITDA = GAAP operating income + D&A (not HL's non-GAAP Adjusted EBITDA).",
 "",
 "Tabs:  Operating Model (annual)  ·  Quarterly Detail (the 20-quarter audit trail).",
 "",
 "Built from public filings for financial-analysis demonstration. Not investment advice.",
]
for i,t in enumerate(notes):
    cov.cell(row=7+i,column=2,value=t).font=f_lbl if not t.startswith("Tie-outs") else f_lblb

# ============================== OPERATING MODEL (ANNUAL) ==============================
om=sheet(wb,"Operating Model")
om.column_dimensions["A"].width=2
om.column_dimensions["B"].width=34
for j in range(len(FYS)): om.column_dimensions[get_column_letter(3+j)].width=12
FCOL=lambda j:get_column_letter(3+j)  # data columns start at C
# title + column header
t=om.cell(row=2,column=2,value="Operating Model: $ in millions unless noted"); t.font=f_h1
hr=4
om.cell(row=hr,column=2,value="Fiscal year (ended Mar 31)").font=f_colh
om.cell(row=hr,column=2).fill=fill_navy; om.cell(row=hr,column=2).alignment=L
for j,fy in enumerate(FYS):
    c=om.cell(row=hr,column=3+j,value=f"FY{fy}"); c.font=f_colh; c.fill=fill_navy; c.alignment=R

row=hr+1
rowmap={}
def sec(title):
    global row
    c=om.cell(row=row,column=2,value=title); c.font=f_secth; c.fill=fill_sec
    for j in range(len(FYS)): om.cell(row=row,column=3+j).fill=fill_sec
    row+=1
def dline(key,label,series,fmt="#,##0.0",bold=False,indent=0):
    """values pulled from SEC (blue)."""
    global row
    c=om.cell(row=row,column=2,value=("   "*indent)+label); c.font=f_lblb if bold else f_lbl
    for j,fy in enumerate(FYS):
        cc=om.cell(row=row,column=3+j,value=series[s(fy)]); cc.font=f_input; cc.number_format=fmt; cc.alignment=R
    rowmap[key]=row; row+=1
def fline(key,label,formula_fn,fmt="0.0%",bold=False):
    """derived, live formula (black)."""
    global row
    c=om.cell(row=row,column=2,value=label); c.font=f_lblb if bold else f_lbl
    for j in range(len(FYS)):
        col=FCOL(j); cc=om.cell(row=row,column=3+j,value=formula_fn(col)); cc.font=f_formu; cc.number_format=fmt; cc.alignment=R
    rowmap[key]=row; row+=1
def blank():
    global row; row+=1

sec("REVENUE")
dline("rev","Total revenue",D["revenue_total"],bold=True)
blank()
sec("SEGMENT REVENUE")
dline("cf","Corporate Finance",D["segment_revenue"]["CF"])
dline("fr","Financial Restructuring",D["segment_revenue"]["FR"])
dline("fva","Financial & Valuation Advisory",D["segment_revenue"]["FVA"])
fline("segtot","Total (check = revenue)",lambda col:f"={col}{rowmap['cf']}+{col}{rowmap['fr']}+{col}{rowmap['fva']}","#,##0.0")
blank()
sec("PROFITABILITY")
dline("ebit","Operating income (EBIT)",D["ebit"])
dline("ebitda","EBITDA (EBIT + D&A)",D["ebitda"])
dline("ni","Net income to HLI",D["ni_to_hli"],bold=True)
fline("opm","Operating margin",lambda col:f"={col}{rowmap['ebit']}/{col}{rowmap['rev']}")
fline("ebm","EBITDA margin",lambda col:f"={col}{rowmap['ebitda']}/{col}{rowmap['rev']}")
fline("nm","Net margin",lambda col:f"={col}{rowmap['ni']}/{col}{rowmap['rev']}")
blank()
sec("COMPENSATION")
dline("comp","Compensation & benefits",D["comp"])
fline("compr","Comp ratio (comp / revenue)",lambda col:f"={col}{rowmap['comp']}/{col}{rowmap['rev']}")
blank()
sec("PEOPLE & ACTIVITY")
dline("md","Managing Directors (year-end)",{k:int(v) for k,v in D["md"]["total"].items()},fmt="#,##0")
fline("rpm","Revenue per MD ($mm)",lambda col:f"={col}{rowmap['rev']}/{col}{rowmap['md']}","#,##0.00")
dline("dcf","Corporate Finance deals closed",{k:int(v) for k,v in D["deals"]["CF"].items()},fmt="#,##0")
dline("dfr","Restructuring deals closed",{k:int(v) for k,v in D["deals"]["FR"].items()},fmt="#,##0")
# footer note
row+=1
n=om.cell(row=row,column=2,value="Blue = pulled from SEC filings.  Black = live formula.  Source: SEC EDGAR CIK 0001302215.")
n.font=Font(name="Calibri",size=8,italic=True,color=GREY)

# ============================== QUARTERLY DETAIL ==============================
qd=sheet(wb,"Quarterly Detail")
qd.column_dimensions["A"].width=2
qd.column_dimensions["B"].width=30
QCOLS=[f"FY{fy}Q{q}" for fy in FYS for q in (1,2,3,4)]
for j in range(len(QCOLS)): qd.column_dimensions[get_column_letter(3+j)].width=9.5
qd.cell(row=2,column=2,value="Quarterly Detail: the audit trail (annual columns above sum these)").font=f_h1
qd.freeze_panes="C6"
hr=5
qd.cell(row=hr,column=2,value="$ in millions / counts").font=f_colh
qd.cell(row=hr,column=2).fill=fill_navy; qd.cell(row=hr,column=2).alignment=L
for j,q in enumerate(QCOLS):
    c=qd.cell(row=hr,column=3+j,value=q.replace("FY","FY ")); c.font=f_colh; c.fill=fill_navy; c.alignment=R
    c.number_format="@"
qrow=hr+1
def qsec(title):
    global qrow
    c=qd.cell(row=qrow,column=2,value=title); c.font=f_secth; c.fill=fill_sec
    for j in range(len(QCOLS)): qd.cell(row=qrow,column=3+j).fill=fill_sec
    qrow+=1
def qline(label,fn,fmt="#,##0.0"):
    global qrow
    qd.cell(row=qrow,column=2,value=label).font=f_lbl
    for j,q in enumerate(QCOLS):
        v=fn(q); cc=qd.cell(row=qrow,column=3+j,value=v); cc.font=f_input; cc.number_format=fmt; cc.alignment=R
    qrow+=1
def segrev(seg): return lambda q: round(QS[q]["seg"][seg]["rev"]/1000,3)
def segmd(seg):  return lambda q: int(QS[q]["seg"][seg]["mds"])
def segdl(seg):  return lambda q: int(QS[q]["seg"][seg]["deals"])
def qfin(line):
    return lambda q: QF[line].get(f"{q[2:6]}-{q[7]}")  # FYyyyyQq -> "yyyy-q"

qsec("SEGMENT REVENUE ($mm)")
qline("Corporate Finance",segrev("CF")); qline("Financial Restructuring",segrev("FR")); qline("Financial & Valuation Advisory",segrev("FVA"))
qline("Total revenue (XBRL)",qfin("rev"))
qsec("MANAGING DIRECTORS (period-end)")
qline("Corporate Finance",segmd("CF"),"#,##0"); qline("Financial Restructuring",segmd("FR"),"#,##0"); qline("Financial & Valuation Advisory",segmd("FVA"),"#,##0")
qsec("DEALS")
qline("Corporate Finance, closed transactions",segdl("CF"),"#,##0"); qline("Financial Restructuring, closed transactions",segdl("FR"),"#,##0")

OUT=os.path.join(PROJ,"Houlihan Lokey - Financial Model.xlsx")
wb.save(OUT)
print("saved",os.path.basename(OUT),", tabs:",wb.sheetnames)
